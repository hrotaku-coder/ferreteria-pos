import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
from tkinter import messagebox
import csv

import productos
from utils import ruta_recurso

class VentanaProductos:
    
    def __init__ (self, ventana):
        self.ventana = ventana
        self.ventana.title("Gestion de Productos")
        self.ventana.configure(bg="#D1D3D5")
        
        # --- CENTRAR VENTANA DE PRODUCTOS ---
        ancho_ventana = 1080
        alto_ventana = 800
        
        ancho_pantalla = self.ventana.winfo_screenwidth()
        alto_pantalla = self.ventana.winfo_screenheight()
        
        x = int((ancho_pantalla / 2) - (ancho_ventana / 2))
        y = int((alto_pantalla / 2) - (alto_ventana / 2))
        
        self.ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

        self.frame_operaciones = tk.LabelFrame(self.ventana, text="Gestion de Productos", font="Arial 12 bold", bg="#D1D3D5")
        self.frame_operaciones.pack(fill="x", padx=10, pady=10,)
        
        self.lbl_buscarproducto = tk.Label(self.frame_operaciones, text="Buscar Producto:", font="Arial 12", bg="#D1D3D5")
        self.lbl_buscarproducto.grid(row=0, column=0, padx=10, pady=10, sticky="e")
        
        self.combo_buscarproducto = ttk.Combobox(self.frame_operaciones, font="sans 12", width=40)
        self.combo_buscarproducto.grid(row=0, column=1, columnspan=2, padx=10, pady=10, sticky="w")

        self.combo_buscarproducto.bind("<KeyRelease>", self.filtrar_productos)
        
        
        self.btn_crearproducto = tk.Button(
            self.frame_operaciones,
            text="Crear Nuevo Producto",
            font=("Arial", 11),
            bg="#D1D3D5",
            width=20,
            command=self.abrir_formulario_producto
            
        )

        self.btn_crearproducto.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        
        
        self.btn_editarproducto = tk.Button(
            self.frame_operaciones,
            text="Editar Producto",
            font=("Arial", 11),
            bg="#D1D3D5",
            width=20,
            command=self.editar_producto
            
        )

        self.btn_editarproducto.grid(row=1, column=1, padx=10, pady=10, sticky="w")
        
        ruta_impot = ruta_recurso("iconos/import.png")
        img = Image.open(ruta_impot)
        img = img.resize((40, 24))
        self.icono_import = ImageTk.PhotoImage(img)
        
        
        self.btn_importarproducto = tk.Button(
            self.frame_operaciones,
            text="  Importar Productos",
            image=self.icono_import,
            compound="left",
            font=("Arial", 11),
            bg="#D1D3D5",
            width=180,
            command=self.importar_productos
            
        )

        self.btn_importarproducto.grid(row=1, column=2, padx=10, pady=10, sticky="w")
        
        self.btn_importarproducto.image = self.icono_import
        
        self.btn_eliminarproducto = tk.Button(
            self.frame_operaciones,
            text="Eliminar Producto",
            font=("Arial", 11),
            bg="#D1D3D5",
            width=20,
            command=self.eliminar_producto
            
        )
        
        self.btn_eliminarproducto.grid(row=1, column=3, padx=10, pady=10, sticky="w")
        
        self.btn_exportarproducto = tk.Button(
            self.frame_operaciones,
            text="Exportar a Excel",
            font=("Arial", 11),
            bg="#D1D3D5",
            width=20,
            command=self.exportar_productos
        )
        self.btn_exportarproducto.grid(row=1, column=4, padx=10, pady=10, sticky="w")
        
        
        self.frame_tabla = tk.LabelFrame(
            self.ventana,
            text="Lista de Productos",
            font="Arial 12 bold",
            bg="#D1D3D5"
        )

        self.frame_tabla.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tabla = ttk.Treeview(self.frame_tabla)

        self.tabla["columns"] = ("referencia", "nombre", "precio1", "precio2", "stock")

        self.tabla.column("#0", width=0, stretch=tk.NO)

        self.tabla.column("referencia", anchor="w", width=150)
        self.tabla.column("nombre", anchor="w", width=300)
        self.tabla.column("precio1", anchor="e", width=100)
        self.tabla.column("precio2", anchor="e", width=100)
        self.tabla.column("stock", anchor="center", width=100)

        self.tabla.heading("#0", text="")
        self.tabla.heading("referencia", text="Referencia")
        self.tabla.heading("nombre", text="Nombre del Producto")
        self.tabla.heading("precio1", text="Precio 1")
        self.tabla.heading("precio2", text="Precio 2")
        self.tabla.heading("stock", text="Stock")

        self.tabla.pack(fill="both", expand=True)
        
        self.cargar_productos()
        
    def abrir_formulario_producto(self):
        self.ventana_form = tk.Toplevel(self.ventana)
        self.ventana_form.title("Nuevo Producto")
        self.ventana_form.configure(bg="#D1D3D5")
        
        ancho_ventana = 600
        alto_ventana = 300
        
        ancho_pantalla = self.ventana_form.winfo_screenwidth() # --- Obtener ancho de pantalla
        alto_pantalla = self.ventana_form.winfo_screenheight() # --- Obtener alto de pantalla
        
        x = int((ancho_pantalla / 2) - (ancho_ventana / 2)) # --- Calcular coordenada x
        y = int((alto_pantalla / 2) - (alto_ventana / 2)) # --- Calcular coordenada y
        
        self.ventana_form.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}") # --- Aplicar tamaño y posición
        
        # convertir la ventana en modal
        self.ventana_form.transient(self.ventana) # --- Hacer que la ventana hija dependa de la ventana principal
        self.ventana_form.grab_set() # --- Capturar todos los eventos en la ventana hija
        self.ventana.focus_set() # --- Establecer el foco en la ventana principal
        
        
        tk.Label(self.ventana_form, text="Refernecia", font="Arial 12", bg="#D1D3D5").place(x=15, y=20)
        self.entry_referen = tk.Entry(self.ventana_form, font="sans 12", width=40)
        self.entry_referen.place(x=15, y=40)
        
        tk.Label(self.ventana_form, text="Nombre del Producto", font="Arial 12", bg="#D1D3D5").place(x=15, y=70)
        self.entry_producto = tk.Entry(self.ventana_form, font="sans 12", width=40)
        self.entry_producto.place(x=15, y=90)
        
        tk.Label(self.ventana_form, text="Precio 1", font="Arial 12", bg="#D1D3D5").place(x=15, y=120)
        self.entry_precio1 = tk.Entry(self.ventana_form, font="sans 12", width=40)
        self.entry_precio1.place(x=15, y=140)
        
        tk.Label(self.ventana_form, text="Precio 2", font="Arial 12", bg="#D1D3D5").place(x=15, y=170)
        self.entry_precio2 = tk.Entry(self.ventana_form, font="sans 12", width=40)
        self.entry_precio2.place(x=15, y=190)
        
        tk.Label(self.ventana_form, text="Stock", font="Arial 12", bg="#D1D3D5").place(x=15, y=220)
        self.entry_stock = tk.Entry(self.ventana_form, font="sans 12", width=40)
        self.entry_stock.place(x=15, y=240)
        
        self.btn_guardar = tk.Button(
            self.ventana_form,
            text="GUARDAR",
            font="Arial 12 bold",
            bg="#D1D3D5",
            width=10,
            command=self.guardar_producto
        )
        
        self.btn_guardar.place(x=450,y=125)
        
    def guardar_producto(self):
        referencia = self.entry_referen.get()
        nombre = self.entry_producto.get()
        precio1 = self.entry_precio1.get()
        precio2 = self.entry_precio2.get()
        stock = self.entry_stock.get()
        
        if not referencia or not nombre or not precio1 or not precio2 or not stock: #-- Validar que no haya campos vacíos
            messagebox.showwarning("Campos vacíos", "Por favor, completa todos los campos")
            return
    
        try:                          #-- Validar que los precios sean números válidos y el stock sea un entero  
            precio1 = float(precio1)
            precio2 = float(precio2)
            stock = int(stock)
        except ValueError:
            messagebox.showerror("Error", "Por favor, ingresa valores numéricos válidos para los precios y el stock")
            return
        
        if hasattr(self, "modo_edicion") and self.modo_edicion:     #-- Si estamos en modo edición, actualizar el producto existente
            productos.actualizar_producto(self.referencia_original, nombre, referencia, precio1, precio2, stock)
            messagebox.showinfo("Producto actualizado", "El producto ha sido actualizado correctamente")
            self.modo_edicion = False
        else:
            productos.agregar_producto(nombre, referencia, precio1, precio2, stock)
            messagebox.showinfo("Producto agregado", "El producto ha sido agregado correctamente")

        self.cargar_productos()
        self.ventana_form.destroy()
        
    def cargar_productos(self):

        for item in self.tabla.get_children():
            self.tabla.delete(item)

        lista = productos.listar_productos()

        for p in lista:
            id_, nombre, referencia, precio1, precio2, stock = p

            self.tabla.insert("", "end", values=(
                referencia,
                nombre,
                precio1,
                precio2,
                stock
            ))
            

    def obtener_producto_seleccionado(self):
        seleccion = self.tabla.focus()

        if not seleccion:
            print("⚠️ Selecciona un producto")
            return None

        datos = self.tabla.item(seleccion, "values")
        return datos
    
    def editar_producto(self):
        datos = self.obtener_producto_seleccionado()

        if not datos:
            return

        referencia, nombre, precio1, precio2, stock = datos

        self.abrir_formulario_producto()

        self.entry_referen.insert(0, referencia)
        self.entry_producto.insert(0, nombre)
        self.entry_precio1.insert(0, precio1)
        self.entry_precio2.insert(0, precio2)
        self.entry_stock.insert(0, stock)

        self.modo_edicion = True
        self.referencia_original = referencia

    def importar_productos(self):
        # 1. Abrir la ventana para elegir solo archivos Excel
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
        )

        if not ruta:
            return # El usuario canceló la ventana

        # 2. Leer el archivo Excel directamente
        try:
            libro = load_workbook(ruta)
            hoja = libro.active

            # Leer desde la fila 2 para saltarse tus encabezados
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                # Validar que la fila no esté completamente vacía
                if fila[0] is None: 
                    continue
                    
                try:
                    # Ajustado a tu orden exacto: Referencia(0), Nombre(1), Precio1(2), Precio2(3), Stock(4)
                    referencia = str(fila[0])
                    nombre = str(fila[1])
                    precio1 = float(fila[2])
                    precio2 = float(fila[3])
                    stock = int(fila[4])

                    productos.agregar_producto(nombre, referencia, precio1, precio2, stock)
                except Exception as e:
                    print("Fila ignorada por formato incorrecto:", fila)

            # 3. Actualizar la tabla visual y avisar
            self.cargar_productos()
            messagebox.showinfo("Importación completada", "Los productos han sido importados correctamente desde tu Excel.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo Excel: {e}")
            
    def exportar_productos(self):
        # 1. Preguntar dónde y con qué nombre guardar el archivo
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Inventario_Ferreteria.xlsx",
            title="Guardar inventario como...",
            filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
        )

        if not ruta:
            return  # Si el usuario cierra la ventana de guardado, no hace nada

        try:
            # 2. Crear un archivo Excel en blanco
            libro = Workbook()
            hoja = libro.active
            hoja.title = "Inventario"

            # 3. Poner los títulos de las columnas en la primera fila
            encabezados = ["Referencia", "Nombre", "Precio 1", "Precio 2", "Stock"]
            hoja.append(encabezados)

            # 4. Traer todos los productos de la base de datos y escribirlos fila por fila
            lista = productos.listar_productos()
            for p in lista:
                id_, nombre, referencia, precio1, precio2, stock = p
                # Se respeta el mismo orden de los encabezados
                hoja.append([referencia, nombre, precio1, precio2, stock])

            # 5. Guardar físicamente el archivo en la ruta elegida
            libro.save(ruta)
            messagebox.showinfo("Exportación exitosa", "El inventario se ha exportado a Excel correctamente.")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un problema al exportar: {e}")

    def filtrar_productos(self, event):
        texto = self.combo_buscarproducto.get().lower()

        for item in self.tabla.get_children():
            self.tabla.delete(item)

        lista = productos.listar_productos()

        for p in lista:
            id_, nombre, referencia, precio1, precio2, stock = p

            if texto in nombre.lower() or texto in referencia.lower():
                self.tabla.insert("", "end", values=(
                    referencia,
                    nombre,
                    precio1,
                    precio2,
                    stock
                ))
                    
    def eliminar_producto(self):
        datos = self.obtener_producto_seleccionado()

        if not datos:
            return

        referencia = datos[0]

        respuesta = tk.messagebox.askyesno(
            "Confirmar",
            f"¿Eliminar producto {referencia}?"
        )

        if respuesta:
            conn = productos.conectar()
            cursor = conn.cursor()

            cursor.execute("DELETE FROM productos WHERE referencia = ?", (referencia,))
            conn.commit()
            conn.close()

            print("🗑 Producto eliminado")

            self.cargar_productos()


            
#if __name__ == "__main__":
#    root = tk.Tk()
#    app = VentanaProductos(root)
#    root.mainloop()